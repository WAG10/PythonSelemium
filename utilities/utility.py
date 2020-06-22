import pytest
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
import logging
import openpyxl


@pytest.mark.usefixtures("setup")
class baseclass:

    def wait_until_link_text(self, text):
        wait = WebDriverWait(self.driver, 15)
        wait.until(expected_conditions.presence_of_element_located((By.LINK_TEXT, text)))
        self.driver.find_element_by_link_text(text).click()  #text = 'India'

    def selectoption_By_visible_text(self,locator,text):
        select_gender = Select(locator) #self.driver.find_element_by_id("exampleFormControlSelect1")
        select_gender.select_by_visible_text(text)

    def get_logger(self):
        logger = logging.getLogger(__name__)  # To get the current file name
        file_handler = logging.FileHandler("mylogfile.log")  # Creating log file
        formatter = logging.Formatter(
            "%(asctime)s : %(levelname)s : %(name)s : %(message)s")  # setting the fomat of the log content
        # 2020-05-31 17:08:50,768 : WARNING : pyTestDemo.test_logging : this prints all warnings
        file_handler.setFormatter(formatter)  # attaching fomatter to filehandler
        logger.addHandler(file_handler)  # attaching the created file to logger handle
        logger.setLevel(logging.INFO)  # Will skip debug and pint from info level and up
        return logger

    @staticmethod
    def get_data_from_excel(testCaseName):
        myWorkbook = openpyxl.load_workbook("E:\\Selenium\\PythonExcelDataSheet.xlsx")
        mysheet = myWorkbook.active
        dict = {}
        for i in range(1, mysheet.max_row + 1):
            for j in range(2, mysheet.max_column + 1):
                if mysheet.cell(row=i, column=1).value == testCaseName:
                    dict[mysheet.cell(row=1, column=j).value] = mysheet.cell(row=i, column=j).value
        return [dict]

