import openpyxl

myWorkbook = openpyxl.load_workbook("E:\\Selenium\\PythonExcelDataSheet.xlsx")

mysheet  = myWorkbook.active

dict = {}

for i in range(1,mysheet.max_row+1):
    for j in range(2,mysheet.max_column+1):
        if mysheet.cell(row=i,column=1).value == "test1":
            dict[mysheet.cell(row=1,column=j).value]= mysheet.cell(row=i,column=j).value

print(dict)



