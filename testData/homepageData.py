import openpyxl


class homepage_data:

    homepage_data_details = {"name": "dhiraj", "email": "xx@gmail.com","Gender": "Male"}, {"name":"shashi","email": "qq@gmail.com","Gender":"Female"}

    @staticmethod
    def get_data_from_excel(testCaseName):
        myWorkbook = openpyxl.load_workbook("E:\\Selenium\\PythonExcelDataSheet.xlsx")
        mysheet = myWorkbook.active
        dict = {}
        for i in range(1, mysheet.max_row + 1):
            for j in range(2, mysheet.max_column + 1):
                if mysheet.cell(row=i, column=1).value == testCaseName:
                    dict[mysheet.cell(row=1, column=j).value] = mysheet.cell(row=i, column=j).value
        return [dict]

